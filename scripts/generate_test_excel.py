#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
创建测试问题的 Excel 文件，然后生成 SQL 并保存结果
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from datetime import datetime

# 配置
BACKEND_BASE_URL = "http://localhost:8080/api"
TIMEOUT = 15

# 测试问题（前5个）
TEST_QUESTIONS = [
    "查询COMPAL产品的出货量",
    "统计上半年的销售总额",
    "显示销售额最高的前10个产品",
    "按品牌统计出货量",
    "查询2024年第一季度的数据"
]

def generate_sql(question: str):
    """调用后端生成 SQL"""
    try:
        response = requests.post(
            f"{BACKEND_BASE_URL}/chat/text2sql",
            json={"message": question},
            timeout=TIMEOUT
        )
        
        if response.status_code == 200:
            data = response.json()
            # 从 tags 中提取 SQL
            sql = ""
            if data.get("tags"):
                for tag in data["tags"]:
                    if tag.get("type") == "sql":
                        sql = tag.get("content", "")
                        break
            
            return {
                "success": True,
                "sql": sql,
                "error": None
            }
        else:
            return {
                "success": False,
                "sql": "",
                "error": f"HTTP {response.status_code}"
            }
    except Exception as e:
        return {
            "success": False,
            "sql": "",
            "error": str(e)
        }

def create_result_excel(questions: list, results: list, output_file: str):
    """创建结果 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SQL 检查"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    success_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    success_font = Font(bold=True, color="FFFFFF", size=11)
    
    error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    error_font = Font(bold=True, color="FFFFFF", size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    # 创建表头
    headers = ["序号", "问题", "生成的 SQL", "状态"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 填充数据
    for idx, (question, result) in enumerate(zip(questions, results), start=1):
        row = idx + 1
        
        # 序号
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = idx
        cell_a.alignment = Alignment(horizontal="center", vertical="top")
        cell_a.border = border
        cell_a.font = Font(size=11)
        
        # 问题
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = question
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_b.border = border
        cell_b.font = Font(size=11)
        
        # SQL
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = result.get("sql", "")
        cell_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_c.border = border
        cell_c.font = Font(size=10)
        
        # 状态
        cell_d = ws.cell(row=row, column=4)
        if result.get("success"):
            cell_d.value = "✓"
            cell_d.fill = success_fill
            cell_d.font = success_font
        else:
            cell_d.value = "✗"
            cell_d.fill = error_fill
            cell_d.font = error_font
        cell_d.alignment = Alignment(horizontal="center", vertical="center")
        cell_d.border = border
        
        # 设置行高
        ws.row_dimensions[row].height = 60
    
    # 设置表头行高
    ws.row_dimensions[1].height = 30
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 保存文件
    wb.save(output_file)
    print(f"[OK] 结果已保存到: {output_file}")

def main():
    print("=" * 70)
    print("Text2SQL 测试 - 前 5 个问题")
    print("=" * 70)
    
    results = []
    for idx, question in enumerate(TEST_QUESTIONS, start=1):
        print(f"\n[{idx}/{len(TEST_QUESTIONS)}] 处理: {question}")
        result = generate_sql(question)
        results.append(result)
        
        status = "[OK]" if result["success"] else "[FAIL]"
        print(f"     Status: {status}")
        if result["sql"]:
            sql_preview = result["sql"][:80] + "..." if len(result["sql"]) > 80 else result["sql"]
            print(f"     SQL: {sql_preview}")
        if result["error"]:
            print(f"     错误: {result['error']}")
    
    # 创建输出文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"text2sql_results_preview_{timestamp}.xlsx"
    create_result_excel(TEST_QUESTIONS, results, output_file)
    
    # 打印统计
    success_count = sum(1 for r in results if r["success"])
    print(f"\n{'=' * 70}")
    print(f"Test completed!")
    print(f"Success: {success_count}/{len(TEST_QUESTIONS)}")
    print(f"Result file: {output_file}")
    print(f"{'=' * 70}")

if __name__ == "__main__":
    main()
