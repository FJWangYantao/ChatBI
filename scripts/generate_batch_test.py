#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
从 Excel 文件中提取问题，批量生成 SQL 结果
支持处理全部问题或仅前 N 个
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from datetime import datetime
import sys

# 配置
BACKEND_BASE_URL = "http://localhost:8080/api"
TIMEOUT = 15

def load_questions_from_excel(excel_file, col_letter='B', limit=None):
    """从 Excel 文件加载问题"""
    questions = []
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 从第2行开始读取（跳过表头）
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f'{col_letter}{row_idx}']
            if cell.value:
                question = str(cell.value).strip()
                if question:
                    questions.append(question)
                    if limit and len(questions) >= limit:
                        break
        
        print(f"[OK] Loaded {len(questions)} questions from column {col_letter}")
        return questions
    except Exception as e:
        print(f"[ERROR] Failed to load Excel: {e}")
        return []

def generate_sql(question: str):
    """Generate SQL via backend API"""
    try:
        response = requests.post(
            f"{BACKEND_BASE_URL}/chat/text2sql",
            json={"message": question},
            timeout=TIMEOUT
        )
        
        if response.status_code == 200:
            data = response.json()
            sql = ""
            if data.get("tags"):
                for tag in data["tags"]:
                    if tag.get("type") == "sql":
                        sql = tag.get("content", "")
                        break
            
            return {
                "success": True,
                "sql": sql,
                "error": None
            }
        else:
            return {
                "success": False,
                "sql": "",
                "error": f"HTTP {response.status_code}"
            }
    except Exception as e:
        return {
            "success": False,
            "sql": "",
            "error": str(e)
        }

def create_result_excel(questions: list, results: list, output_file: str):
    """Create result Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SQL Results"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    success_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    success_font = Font(bold=True, color="FFFFFF")
    
    error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    error_font = Font(bold=True, color="FFFFFF")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    # Create header
    headers = ["No.", "Question", "Generated SQL", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data
    for idx, (question, result) in enumerate(zip(questions, results), start=1):
        row = idx + 1
        
        # No.
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = idx
        cell_a.alignment = Alignment(horizontal="center", vertical="top")
        cell_a.border = border
        
        # Question
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = question
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_b.border = border
        
        # SQL
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = result.get("sql", "")
        cell_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_c.border = border
        
        # Status
        cell_d = ws.cell(row=row, column=4)
        if result.get("success"):
            cell_d.value = "[OK]"
            cell_d.fill = success_fill
            cell_d.font = success_font
        else:
            cell_d.value = "[FAIL]"
            cell_d.fill = error_fill
            cell_d.font = error_font
        cell_d.alignment = Alignment(horizontal="center", vertical="center")
        cell_d.border = border
        
        # Set row height
        ws.row_dimensions[row].height = 60
    
    # Set header row height
    ws.row_dimensions[1].height = 30
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Save file
    wb.save(output_file)
    print(f"[OK] Results saved to: {output_file}")

def main():
    # Get parameters
    if len(sys.argv) < 2:
        print("Usage: python generate_batch_test.py <source_excel> [limit]")
        print("Example: python generate_batch_test.py source.xlsx 5")
        print("         python generate_batch_test.py source.xlsx    (all)")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
    
    # Load questions
    questions = load_questions_from_excel(excel_file, limit=limit)
    
    if not questions:
        print("[ERROR] No questions found")
        sys.exit(1)
    
    print(f"[OK] Starting test for {len(questions)} questions")
    print("=" * 70)
    
    results = []
    for idx, question in enumerate(questions, start=1):
        print(f"[{idx}/{len(questions)}] {question[:50]}...")
        result = generate_sql(question)
        results.append(result)
        
        if result["success"]:
            sql_preview = result["sql"][:60] + "..." if len(result["sql"]) > 60 else result["sql"]
            print(f"        Status: [OK] - {sql_preview}")
        else:
            print(f"        Status: [FAIL] - {result['error']}")
    
    # Create output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    limit_str = f"_top{limit}" if limit else ""
    output_file = f"text2sql_batch_results{limit_str}_{timestamp}.xlsx"
    create_result_excel(questions, results, output_file)
    
    # Print statistics
    success_count = sum(1 for r in results if r["success"])
    success_rate = 100 * success_count // len(questions) if questions else 0
    print(f"\n{'=' * 70}")
    print(f"Test completed!")
    print(f"Total: {len(questions)}")
    print(f"Success: {success_count}/{len(questions)} ({success_rate}%)")
    print(f"Result file: {output_file}")
    print(f"{'=' * 70}")

if __name__ == "__main__":
    main()
