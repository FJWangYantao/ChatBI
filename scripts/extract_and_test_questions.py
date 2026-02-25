#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
从 Excel 文件中提取问题，生成 SQL，并保存到新的 Excel 文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
import sys
from datetime import datetime
import time

# 配置
BACKEND_BASE_URL = "http://localhost:8080/api"
TIMEOUT = 15

def load_questions_from_excel(excel_file, col_letter='B'):
    """从 Excel 文件的指定列加载问题"""
    questions = []
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 从第2行开始读取（跳过表头）
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[f'{col_letter}{row_idx}']
            if cell.value:
                question = str(cell.value).strip()
                if question:
                    questions.append(question)
        
        print(f"[OK] 从 {excel_file} 的 {col_letter} 列加载了 {len(questions)} 个问题")
        return questions
    except Exception as e:
        print(f"[ERROR] 加载 Excel 文件失败: {e}")
        return []

def generate_sql(question: str):
    """调用后端生成 SQL - 使用新的 /api/chat/message 接口支持 NER"""
    start_time = time.time()
    try:
        # 使用新的 /api/chat/message 接口，支持 NER 增强
        response = requests.post(
            f"{BACKEND_BASE_URL}/chat/message",
            json={"message": question, "conversationId": None},
            timeout=TIMEOUT
        )
        
        elapsed_time = time.time() - start_time
        
        if response.status_code == 200:
            data = response.json()
            # 从 tags 中提取 SQL
            sql = ""
            if data.get("tags"):
                for tag in data["tags"]:
                    if tag.get("type") == "sql":
                        sql = tag.get("content", "")
                        break
            
            return {
                "success": True,
                "sql": sql,
                "message": data.get("reply", ""),
                "error": None,
                "elapsed_time": elapsed_time,
                "conversation_id": data.get("conversationId")
            }
        else:
            return {
                "success": False,
                "sql": "",
                "message": "",
                "error": f"HTTP {response.status_code}",
                "elapsed_time": elapsed_time,
                "conversation_id": None
            }
    except Exception as e:
        elapsed_time = time.time() - start_time
        return {
            "success": False,
            "sql": "",
            "message": "",
            "error": str(e),
            "elapsed_time": elapsed_time,
            "conversation_id": None
        }

def create_result_excel(questions: list, results: list, output_file: str):
    """创建结果 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试结果"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    success_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    success_font = Font(bold=True, color="FFFFFF")
    
    error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    error_font = Font(bold=True, color="FFFFFF")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 5   # 序号
    ws.column_dimensions['B'].width = 30  # 问题
    ws.column_dimensions['C'].width = 50  # SQL
    ws.column_dimensions['D'].width = 12  # 状态
    ws.column_dimensions['E'].width = 12  # 耗时（秒）
    
    # 创建表头
    headers = ["序号", "问题", "生成的 SQL", "状态", "耗时（秒）"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 填充数据
    for idx, (question, result) in enumerate(zip(questions, results), start=1):
        row = idx + 1
        
        # 序号
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = idx
        cell_a.alignment = Alignment(horizontal="center", vertical="top")
        cell_a.border = border
        
        # 问题
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = question
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_b.border = border
        
        # SQL
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = result.get("sql", "")
        cell_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_c.border = border
        
        # 状态
        cell_d = ws.cell(row=row, column=4)
        if result.get("success"):
            cell_d.value = "✓ 成功"
            cell_d.fill = success_fill
            cell_d.font = success_font
        else:
            cell_d.value = "✗ 失败"
            cell_d.fill = error_fill
            cell_d.font = error_font
        cell_d.alignment = Alignment(horizontal="center", vertical="center")
        cell_d.border = border
        
        # 耗时（秒）
        cell_e = ws.cell(row=row, column=5)
        elapsed_time = result.get("elapsed_time", 0)
        cell_e.value = f"{elapsed_time:.2f}"
        cell_e.alignment = Alignment(horizontal="center", vertical="center")
        cell_e.border = border
        
        # 设置行高
        ws.row_dimensions[row].height = max(20, len(question) // 20 * 20)
    
    # 设置表头行高
    ws.row_dimensions[1].height = 25
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 保存文件
    wb.save(output_file)
    print(f"[OK] 结果已保存到: {output_file}")

def main():
    # 获取源 Excel 文件
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = "sql_functionality_test_results.xlsx"
    
    # 加载问题
    questions = load_questions_from_excel(excel_file)
    
    if not questions:
        print("[ERROR] 没有找到任何问题")
        sys.exit(1)
    
    # 只取前5个
    test_questions = questions
    print(f"\n[OK] 开始测试前 {len(test_questions)} 个问题")
    print("=" * 80)
    
    results = []
    for idx, question in enumerate(test_questions, start=1):
        print(f"\n[{idx}/{len(test_questions)}] 处理: {question}")
        result = generate_sql(question)
        results.append(result)
        
        status = "✓ 成功" if result["success"] else "✗ 失败"
        elapsed_time = result.get("elapsed_time", 0)
        print(f"     状态: {status}")
        print(f"     耗时: {elapsed_time:.2f} 秒")
        if result["sql"]:
            sql_preview = result["sql"][:100] + "..." if len(result["sql"]) > 100 else result["sql"]
            print(f"     SQL: {sql_preview}")
        if result["error"]:
            print(f"     错误: {result['error']}")
    
    # 创建输出文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"text2sql_results_{timestamp}.xlsx"
    create_result_excel(test_questions, results, output_file)
    
    # 打印统计
    success_count = sum(1 for r in results if r["success"])
    total_time = sum(r.get("elapsed_time", 0) for r in results)
    avg_time = total_time / len(results) if results else 0
    
    print(f"\n{'=' * 80}")
    print(f"测试完成！")
    print(f"  成功率: {success_count}/{len(test_questions)} ({100*success_count//len(test_questions)}%)")
    print(f"  总耗时: {total_time:.2f} 秒")
    print(f"  平均耗时: {avg_time:.2f} 秒/条")
    print(f"  结果文件: {output_file}")

if __name__ == "__main__":
    main()
